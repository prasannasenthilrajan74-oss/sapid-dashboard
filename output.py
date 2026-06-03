
#!/usr/bin/env python3
import os
import sys
import csv
import json
import math
import random
import datetime
import threading
import urllib.request
import urllib.error
import urllib.parse
import webbrowser
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# Check for openpyxl to support Excel (.xlsx) parsing natively on desktop
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─── Configuration & Paths ───────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(SCRIPT_DIR, "sap_license_config.json")
LOG_PATH = os.path.join(SCRIPT_DIR, "sap_license_error.log")

DEFAULT_LIMITS = {
    "AX": 78,
    "AY": 6,
    "FX": 37,
    "HC": 83,
    "HD": 41,
    "Other": 150
}

LICENSE_COLORS = {
    "AX": "#0ea5e9",       # Sky Blue
    "AY": "#6366f1",       # Indigo
    "FX": "#a855f7",       # Purple
    "HC": "#ec4899",       # Pink
    "HD": "#14b8a6",       # Teal
    "Other": "#64748b"     # Muted Slate
}

DEPT_COLORS = {
    "SD": "#f43f5e",
    "QM": "#10b981",
    "PM": "#f59e0b",
    "R&D": "#3b82f6",
    "MM": "#8b5cf6",
    "HCM": "#06b6d4",
    "PP": "#84cc16",
    "MED": "#ec4899",
    "FICO": "#14b8a6",
    "Other": "#64748b"
}

# ─── Robust Column Keys mapping ───────────────────────────────────────────────
SAPID_KEYS = ['SAPID', 'sapid', 'sap id', 'sap_id', 'id', 'user id', 'userid', 'emp id', 'empid', 'employee id']
NAME_KEYS = ['Name', 'name', 'user name', 'username', 'employee name', 'emp name', 'full name', 'fullname']
LICENSE_KEYS = ['License', 'license', 'license type', 'licenseid', 'license id']
DEPT_KEYS = ['Department', 'department', 'dept', 'dep', 'division', 'div']
FUNC_KEYS = ['Function', 'function', 'role', 'job title', 'jobtitle', 'group', 'user group']
LOGON_KEYS = ['Last Logon', 'last logon', 'lastlogon', 'last_logon', 'logon', 'last login', 'lastlogin']
GROUP_KEYS = ['Group', 'group', 'user group', 'usergroup', 'grp']

def get_row_value(row, target_keys, default_val=""):
    if not row:
        return default_val
    for k in target_keys:
        if k in row and row[k] is not None:
            return str(row[k]).strip()
    
    # Try normalized keys
    for actual_key in row.keys():
        norm_key = str(actual_key).lower().replace(" ", "").replace("_", "").replace("-", "")
        for tk_key in target_keys:
            norm_tk = tk_key.lower().replace(" ", "").replace("_", "").replace("-", "")
            if norm_key == norm_tk:
                val = row[actual_key]
                if val is not None:
                    return str(val).strip()
    return default_val

# ─── Logging Helpers ──────────────────────────────────────────────────────────
def append_log(level, message):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    clean_level = str(level).strip().upper()[:20]
    # Remove control chars
    clean_message = "".join(ch for ch in str(message) if not ch.isspace() or ch in [' ', '\t', '\n'])[:2000]
    log_line = f"[{timestamp}] [{clean_level}] {clean_message}\n"
    
    try:
        if os.path.exists(LOG_PATH) and os.path.getsize(LOG_PATH) > 1000000:
            try:
                os.remove(LOG_PATH)
            except Exception:
                pass
        with open(LOG_PATH, "a", encoding="utf-8") as f:
            f.write(log_line)
    except Exception as e:
        print(f"Failed to write log: {e}")

def read_logs():
    if not os.path.exists(LOG_PATH):
        return "No logs found."
    try:
        with open(LOG_PATH, "r", encoding="utf-8") as f:
            lines = f.readlines()
        return "".join(lines[-50:])
    except Exception as e:
        return f"Error reading logs: {e}"

def clear_logs():
    try:
        if os.path.exists(LOG_PATH):
            os.remove(LOG_PATH)
    except Exception as e:
        append_log("ERROR", f"Failed to clear logs: {e}")

# ─── Mock Data Generator ──────────────────────────────────────────────────────
def generate_mock_data():
    first_names = ["Emma", "Lucas", "Aarav", "Yuto", "Oliver", "Chloe", "Noah", "Sophia", "Liam", "Mia", "Alexander", "Charlotte", "Daniel", "Amelia", "Ethan", "Zoe", "Mason", "Harper", "Logan", "Evelyn"]
    last_names = ["Smith", "Weber", "Sharma", "Tanaka", "Taylor", "Dubois", "Jones", "Miller", "Davis", "Garcia", "Rodriguez", "Wilson", "Martinez", "Anderson", "Taylor", "Thomas", "White", "Harris", "Martin", "Clark"]
    
    licenses = ["AX", "AY", "FX", "HC", "HD"]
    license_weights = [0.40, 0.05, 0.10, 0.30, 0.15]
    
    depts = ["QM", "PM", "R&D", "MM", "HCM", "PP", "SD", "FICO", "MED"]
    
    mock_details = {
        "QM": ["Quality Engineer", "QA Specialist", "Inspector", "Audit Lead"],
        "PM": ["Maintenance Tech", "Reliability Eng", "Planner", "Supervisor"],
        "R&D": ["Research Chemist", "Lab Technician", "Product Designer", "R&D Scientist"],
        "MM": ["Materials Handler", "Procurement Specialist", "Buyer", "Inventory Planner"],
        "HCM": ["HR Business Partner", "Talent Acquisition", "Comp Analyst", "HR Associate"],
        "PP": ["Production Planner", "Scheduler", "Shopfloor Operator", "Assembly Lead"],
        "SD": ["Sales Exec", "Customer Manager", "Billing Specialist", "Account Manager"],
        "FICO": ["Financial Analyst", "Cost Accountant", "Tax Analyst", "Internal Auditor"],
        "MED": ["Health Advisor", "Safety Specialist", "Clinic Coordinator", "Occupational Nurse"]
    }
    
    countries = ["US", "DE", "IN", "JP", "UK", "FR"]
    
    data = []
    # Generate 150 users
    for i in range(1, 151):
        dept = random.choice(depts)
        func = random.choice(mock_details[dept])
        lic = random.choices(licenses, weights=license_weights)[0]
        country = random.choice(countries)
        plant_prefix = f"{country}{random.randint(10, 99)}"
        sapid = f"{plant_prefix}{i:03d}"
        name = f"{random.choice(first_names)} {random.choice(last_names)}"
        
        days_ago = random.randint(0, 180)
        benchmark = datetime.date(2026, 6, 2)
        logon_date = benchmark - datetime.timedelta(days=days_ago)
        logon_str = logon_date.strftime("%Y-%m-%d")
        
        group = f"GRP_{dept}_{random.randint(1, 3)}"
        
        data.append({
            "SAPID": sapid,
            "Name": name,
            "License": lic,
            "Department": dept,
            "Function": func,
            "Last Logon": logon_str,
            "Group": group
        })
    return data

# ─── URL Conversion Helper ────────────────────────────────────────────────────
def get_sheets_csv_url(url_str):
    url_str = url_str.strip()
    if not url_str:
        return None
    if "/pub" in url_str:
        if "output=csv" in url_str:
            return url_str
        # Strip query and add output=csv
        if "?" in url_str:
            base_url = url_str.split("?")[0]
        else:
            base_url = url_str
        if base_url.endswith("/pub"):
            return f"{base_url}?output=csv"
        return url_str
        
    # Standard sheet URL matching
    # Pattern: /spreadsheets/d/<ID>/
    if "/spreadsheets/d/" in url_str:
        parts = url_str.split("/spreadsheets/d/")
        if len(parts) > 1:
            sheet_id = parts[1].split("/")[0]
            # Parse gid if present
            gid = "0"
            parsed_query = urllib.parse.urlparse(url_str).query
            query_params = urllib.parse.parse_qs(parsed_query)
            if "gid" in query_params:
                gid = query_params["gid"][0]
            return f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&gid={gid}"
    return url_str

# ─── Help Text Content ────────────────────────────────────────────────────────
USER_MANUAL_TEXT = """SAPID License Analyser - Desktop User Manual

1. Syncing with Google Sheets:
   - Paste a public Google Sheet URL into the top link bar.
   - Example format: https://docs.google.com/spreadsheets/d/...
   - Make sure General Access is set to 'Anyone with the link can view' in Google Sheets.
   - Click 'Sync' (or press Enter). The tool extracts data automatically.

2. Google Apps Script Web App Integration:
   - If your organization blocks public links, deploy a Google Apps Script returning JSON.
   - Apps Script Web App links ending in '/exec' are handled and translated dynamically.

3. Uploading Local Sheets:
   - Click the 'Upload Excel / CSV' area in the header.
   - Select any local .csv or .xlsx (requires openpyxl installed) file.
   - The application parses data inside temporary system memory.

4. Editable License Purchased Limits:
   - In the right-hand License distribution panel, every license type has an editable 'Limit' box.
   - Type in a number and press Enter (or focus out). The Total Purchased and Available counts update instantly.
   - Quotas are saved automatically to your configuration files.

5. Interactive Plant ID Explorer:
   - The third dashboard bar chart groups users by their SAPID Plant ID Prefix (first 4 characters).
   - Click on any bar in the chart to open a popup drilldown modal showing the profile details of all matching users, sorted by department.
"""

PRIVACY_TEXT = """SAPID License Analyser - Privacy Policy

1. Local-First Design:
   All calculations, files, configurations, and logs run locally on your system. None of your data, sheet configurations, or uploaded spreadsheets are sent to external trackers.

2. Sheets Connection:
   Communication happens directly between your local script and Google endpoints. No credentials or sheet rows are routed through intermediate APIs.

3. Local Storage:
   Settings, limits, and cached spreadsheet rows are saved in a local JSON configuration file inside the application directory.

4. App Diagnostic Logs:
   Log files are written locally to assist in system diagnostics and debugging. You can view, open, or purge them at any time.
"""

LICENSES_TEXT = """SAPID License Analyser - Third-Party Open Source Notices

- Python Standard Library (PSF License) - Core Tkinter, JSON, HTTP/URL client modules.
- openpyxl (MIT License) - Used to parse spreadsheet file structures natively on desktop.
- Charting & Components are drawn directly on custom Canvas grids without external GUI packages, ensuring lightweight performance.
"""

# ─── Dialog Modals ────────────────────────────────────────────────────────────
class TextModal(tk.Toplevel):
    def __init__(self, parent, title, heading, text_content):
        super().__init__(parent)
        self.title(title)
        self.geometry("600x500")
        self.configure(bg="#0b0f19")
        self.transient(parent)
        self.grab_set()
        
        # Heading
        lbl_head = tk.Label(self, text=heading, font=("Helvetica", 14, "bold"), bg="#0b0f19", fg="#f8fafc", anchor="w")
        lbl_head.pack(fill="x", padx=20, pady=15)
        
        # Scrolled Text
        frame = tk.Frame(self, bg="#161e31", bd=1, relief="solid", highlightthickness=0)
        frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        txt_area = tk.Text(frame, bg="#161e31", fg="#f8fafc", insertbackground="#f8fafc", font=("Consolas", 10), wrap="word", bd=0, padx=10, pady=10)
        txt_area.insert("1.0", text_content)
        txt_area.config(state="disabled")
        txt_area.pack(side="left", fill="both", expand=True)
        
        scroll = ttk.Scrollbar(frame, orient="vertical", command=txt_area.yview)
        scroll.pack(side="right", fill="y")
        txt_area.config(yscrollcommand=scroll.set)
        
        # Close Button
        btn_close = tk.Button(self, text="Close", font=("Helvetica", 10, "bold"), bg="#334155", fg="#f8fafc", activebackground="#475569", activeforeground="#f8fafc", bd=0, padx=15, pady=6, cursor="hand2", command=self.destroy)
        btn_close.pack(pady=(0, 20))

class UserDetailModal(tk.Toplevel):
    def __init__(self, parent, user_data):
        super().__init__(parent)
        self.title(f"User Profile: {user_data.get('Name', 'Unknown')}")
        self.geometry("550x380")
        self.configure(bg="#0b0f19")
        self.transient(parent)
        self.grab_set()
        
        lbl_title = tk.Label(self, text="SAP Account Profile Details", font=("Helvetica", 14, "bold"), bg="#0b0f19", fg="#0ea5e9", anchor="w")
        lbl_title.pack(fill="x", padx=20, pady=15)
        
        # Grid frame
        frame = tk.Frame(self, bg="#161e31", highlightbackground="#2a354f", highlightthickness=1, padx=20, pady=20)
        frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        fields = [
            ("SAPID:", "SAPID"),
            ("User Name:", "Name"),
            ("License Type:", "License"),
            ("Department:", "Department"),
            ("Function / Role:", "Function"),
            ("Last Logon Date:", "Last Logon"),
            ("User Group:", "Group")
        ]
        
        for idx, (label_text, key) in enumerate(fields):
            lbl_key = tk.Label(frame, text=label_text, font=("Helvetica", 10, "bold"), bg="#161e31", fg="#94a3b8", anchor="w")
            lbl_key.grid(row=idx, column=0, sticky="w", pady=6)
            
            val = user_data.get(key, "-")
            if val == "":
                val = "-"
            
            lbl_val = tk.Label(frame, text=val, font=("Helvetica", 10), bg="#161e31", fg="#f8fafc", anchor="w")
            lbl_val.grid(row=idx, column=1, sticky="w", padx=15, pady=6)
            
            # Badge styles for License/Dept
            if key == "License":
                lbl_val.config(fg="#a5b4fc", font=("Helvetica", 10, "bold"))
            elif key == "Department":
                lbl_val.config(fg="#d8b4fe", font=("Helvetica", 10, "bold"))
            elif key == "Last Logon":
                lbl_val.config(fg="#38bdf8", font=("Helvetica", 10, "bold"))
        
        btn_close = tk.Button(self, text="Close", font=("Helvetica", 10, "bold"), bg="#334155", fg="#f8fafc", activebackground="#475569", activeforeground="#f8fafc", bd=0, padx=15, pady=6, cursor="hand2", command=self.destroy)
        btn_close.pack(pady=(0, 15))

class PlantIDDetailsModal(tk.Toplevel):
    def __init__(self, parent, prefix, matching_users):
        super().__init__(parent)
        self.title(f"Plant ID drilldown: {prefix}")
        self.geometry("800x480")
        self.configure(bg="#0b0f19")
        self.transient(parent)
        self.grab_set()
        
        # Header text
        lbl_head = tk.Label(self, text=f"Plant ID Group: {prefix}", font=("Helvetica", 14, "bold"), bg="#0b0f19", fg="#f8fafc", anchor="w")
        lbl_head.pack(fill="x", padx=20, pady=(15, 2))
        
        lbl_sub = tk.Label(self, text=f"Total Users matching Prefix: {len(matching_users)} (sorted by Department)", font=("Helvetica", 9), bg="#0b0f19", fg="#94a3b8", anchor="w")
        lbl_sub.pack(fill="x", padx=20, pady=(0, 10))
        
        # Frame for table
        frame = tk.Frame(self, bg="#161e31")
        frame.pack(fill="both", expand=True, padx=20, pady=(0, 15))
        
        cols = ("sapid", "name", "license", "department", "function", "lastlogon")
        self.tree = ttk.Treeview(frame, columns=cols, show="headings", selectmode="browse")
        
        self.tree.heading("sapid", text="SAPID")
        self.tree.heading("name", text="User Name")
        self.tree.heading("license", text="License")
        self.tree.heading("department", text="Department")
        self.tree.heading("function", text="Function")
        self.tree.heading("lastlogon", text="Last Logon")
        
        self.tree.column("sapid", width=80, anchor="center")
        self.tree.column("name", width=150, anchor="w")
        self.tree.column("license", width=70, anchor="center")
        self.tree.column("department", width=90, anchor="center")
        self.tree.column("function", width=220, anchor="w")
        self.tree.column("lastlogon", width=100, anchor="center")
        
        # Style treeview inside the dialog
        style = ttk.Style()
        style.configure("Popup.Treeview", background="#1e293b", foreground="#f8fafc", fieldbackground="#1e293b", rowheight=26, font=("Helvetica", 9))
        self.tree.config(style="Popup.Treeview")
        
        self.tree.pack(side="left", fill="both", expand=True)
        
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.tree.config(yscrollcommand=scrollbar.set)
        
        # Sort users by department, then name
        sorted_users = sorted(matching_users, key=lambda x: (x.get("Department", ""), x.get("Name", "")))
        
        for u in sorted_users:
            self.tree.insert("", "end", values=(
                u.get("SAPID", ""),
                u.get("Name", ""),
                u.get("License", ""),
                u.get("Department", ""),
                u.get("Function", ""),
                u.get("Last Logon", "")
            ))
            
        self.tree.bind("<Double-1>", self.on_double_click)
        
        btn_close = tk.Button(self, text="Close", font=("Helvetica", 10, "bold"), bg="#334155", fg="#f8fafc", activebackground="#475569", activeforeground="#f8fafc", bd=0, padx=15, pady=6, cursor="hand2", command=self.destroy)
        btn_close.pack(pady=(0, 15))
        
    def on_double_click(self, event):
        item = self.tree.selection()
        if not item:
            return
        vals = self.tree.item(item[0], "values")
        # Find matching record
        user_id = vals[0]
        # Open profile detail dialog
        for u in main_app.raw_data:
            if u.get("SAPID") == user_id:
                UserDetailModal(self, u)
                break

class ErrorLogsModal(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Local Diagnostic Error Logs")
        self.geometry("700x480")
        self.configure(bg="#0b0f19")
        self.transient(parent)
        self.grab_set()
        
        lbl_title = tk.Label(self, text="System Error Logs", font=("Helvetica", 14, "bold"), bg="#0b0f19", fg="#f43f5e", anchor="w")
        lbl_title.pack(fill="x", padx=20, pady=(15, 2))
        
        lbl_sub = tk.Label(self, text="Last 50 diagnostic events logged by the system application", font=("Helvetica", 9), bg="#0b0f19", fg="#94a3b8", anchor="w")
        lbl_sub.pack(fill="x", padx=20, pady=(0, 10))
        
        # Scrolled Text Box
        frame = tk.Frame(self, bg="#161e31", bd=1, relief="solid", highlightthickness=0)
        frame.pack(fill="both", expand=True, padx=20, pady=(0, 15))
        
        self.txt = tk.Text(frame, bg="#161e31", fg="#f43f5e", insertbackground="#f8fafc", font=("Consolas", 9), wrap="none", bd=0, padx=10, pady=10)
        self.txt.pack(side="left", fill="both", expand=True)
        
        ysb = ttk.Scrollbar(frame, orient="vertical", command=self.txt.yview)
        ysb.pack(side="right", fill="y")
        xsb = ttk.Scrollbar(self, orient="horizontal", command=self.txt.xview)
        xsb.pack(fill="x", padx=20)
        
        self.txt.config(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        
        self.load_logs_text()
        
        # Footer buttons
        btn_frame = tk.Frame(self, bg="#0b0f19")
        btn_frame.pack(fill="x", padx=20, pady=15)
        
        btn_folder = tk.Button(btn_frame, text="📂 Open Logs Location", font=("Helvetica", 9, "bold"), bg="#1e293b", fg="#f8fafc", activebackground="#334155", activeforeground="#f8fafc", bd=0, padx=12, pady=6, cursor="hand2", command=self.open_folder)
        btn_folder.pack(side="left")
        
        btn_clear = tk.Button(btn_frame, text="🗑️ Purge Logs", font=("Helvetica", 9, "bold"), bg="#ef4444", fg="#f8fafc", activebackground="#dc2626", activeforeground="#f8fafc", bd=0, padx=12, pady=6, cursor="hand2", command=self.purge_logs)
        btn_clear.pack(side="left", padx=10)
        
        btn_close = tk.Button(btn_frame, text="Close", font=("Helvetica", 9, "bold"), bg="#334155", fg="#f8fafc", activebackground="#475569", activeforeground="#f8fafc", bd=0, padx=15, pady=6, cursor="hand2", command=self.destroy)
        btn_close.pack(side="right")
        
    def load_logs_text(self):
        self.txt.config(state="normal")
        self.txt.delete("1.0", "end")
        self.txt.insert("1.0", read_logs())
        self.txt.config(state="disabled")
        
    def purge_logs(self):
        if messagebox.askyesno("Confirm Purge", "Are you sure you want to truncate the log history?", parent=self):
            clear_logs()
            self.load_logs_text()
            
    def open_folder(self):
        try:
            # Create if it doesn't exist
            if not os.path.exists(LOG_PATH):
                with open(LOG_PATH, "w", encoding="utf-8") as f:
                    f.write("")
            if sys.platform == "win32":
                os.startfile(os.path.dirname(LOG_PATH))
            elif sys.platform == "darwin":
                import subprocess
                subprocess.Popen(["open", os.path.dirname(LOG_PATH)])
            else:
                import subprocess
                subprocess.Popen(["xdg-open", os.path.dirname(LOG_PATH)])
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open explorer: {e}", parent=self)

# ─── Custom Canvas Tooltip Handler ────────────────────────────────────────────
class CanvasTooltip:
    def __init__(self, canvas):
        self.canvas = canvas
        self.tooltip_rect = None
        self.tooltip_text = None
        
    def show(self, x, y, text):
        self.hide()
        # Draw box above cursor
        pad = 6
        self.tooltip_text = self.canvas.create_text(
            x, y - 20,
            text=text,
            fill="#f8fafc",
            font=("Helvetica", 8, "bold"),
            anchor="center",
            tags="tooltip"
        )
        bbox = self.canvas.bbox(self.tooltip_text)
        if bbox:
            tx1, ty1, tx2, ty2 = bbox
            self.tooltip_rect = self.canvas.create_rectangle(
                tx1 - pad, ty1 - pad,
                tx2 + pad, ty2 + pad,
                fill="#0f172a",
                outline="#0ea5e9",
                width=1,
                tags="tooltip"
            )
            # Raise text above rect
            self.canvas.tag_raise(self.tooltip_text, self.tooltip_rect)
            
    def hide(self):
        self.canvas.delete("tooltip")
        self.tooltip_rect = None
        self.tooltip_text = None

# ─── Custom Card widgets ──────────────────────────────────────────────────────
class KPICard(tk.Frame):
    def __init__(self, parent, title, value, icon, color):
        super().__init__(parent, bg="#161e31", highlightbackground="#2a354f", highlightthickness=1)
        
        # Color line on the left
        color_bar = tk.Frame(self, bg=color, width=4)
        color_bar.pack(side="left", fill="y")
        
        # Info container
        info_frame = tk.Frame(self, bg="#161e31")
        info_frame.pack(side="left", fill="both", expand=True, padx=12, pady=10)
        
        self.lbl_title = tk.Label(info_frame, text=title.upper(), font=("Helvetica", 7, "bold"), bg="#161e31", fg="#94a3b8", anchor="w")
        self.lbl_title.pack(fill="x")
        
        self.lbl_val = tk.Label(info_frame, text=value, font=("Helvetica", 14, "bold"), bg="#161e31", fg="#f8fafc", anchor="w")
        self.lbl_val.pack(fill="x", pady=(2, 0))
        
        # Icon
        self.lbl_icon = tk.Label(self, text=icon, font=("Helvetica", 18), bg="#161e31", fg=color)
        self.lbl_icon.pack(side="right", padx=12)
        
    def update_val(self, val):
        self.lbl_val.config(text=str(val))

# ─── Main Application Class ───────────────────────────────────────────────────
class SAPLicenseAnalyserApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("SAPID license analyser")
        self.geometry("1300x820")
        self.configure(bg="#0b0f19")
        
        self.raw_data = []
        self.filtered_data = []
        self.limits = DEFAULT_LIMITS.copy()
        self.active_url = ""
        self.current_tab = "dashboard"
        
        # Ledger State
        self.ledger_sort_col = "sapid"
        self.ledger_sort_asc = True
        self.ledger_page = 1
        self.ledger_page_size = 10
        
        # Explorer State
        self.explorer_page = 1
        self.explorer_page_size = 6
        self.explorer_selected_func = ""
        
        # Load local configurations
        self.load_config()
        
        # Styling Setup
        self.setup_styles()
        
        # Build UI Elements
        self.build_ui()
        
        # Initial UI Refresh
        self.refresh_ui()
        
    def setup_styles(self):
        self.style = ttk.Style()
        self.style.theme_use('clam')
        
        # Dark styling variables for Ttk
        self.style.configure("TFrame", background="#0b0f19")
        self.style.configure("TLabel", background="#0b0f19", foreground="#f8fafc")
        
        # Scrollbars
        self.style.configure("TScrollbar", background="#1e293b", troughcolor="#0b0f19", bordercolor="#2a354f", arrowcolor="#f8fafc")
        self.style.map("TScrollbar", background=[("active", "#334155")])
        
        # Combobox / OptionMenu styling overrides
        self.style.configure("TCombobox", fieldbackground="#1e293b", background="#1e293b", foreground="#f8fafc", arrowcolor="#f8fafc", bordercolor="#2a354f")
        
        # Treeview styling
        self.style.configure("Treeview", background="#161e31", foreground="#f8fafc", fieldbackground="#161e31", rowheight=26, bordercolor="#2a354f", font=("Helvetica", 9))
        self.style.map("Treeview", background=[("selected", "#0ea5e9")], foreground=[("selected", "#0b0f19")])
        
        self.style.configure("Treeview.Heading", background="#1e293b", foreground="#94a3b8", relief="flat", font=("Helvetica", 9, "bold"))
        self.style.map("Treeview.Heading", background=[("active", "#334155")], foreground=[("active", "#f8fafc")])

    def build_ui(self):
        # Master Wrapper Grid
        self.grid_rowconfigure(0, weight=0)  # Header sync row
        self.grid_rowconfigure(1, weight=0)  # Nav tabs row
        self.grid_rowconfigure(2, weight=1)  # Main pages container
        self.grid_rowconfigure(3, weight=0)  # Footer
        self.grid_columnconfigure(0, weight=1)
        
        # ─── HEADER ROW ───
        header_frame = tk.Frame(self, bg="#0b0f19", padx=20, pady=15)
        header_frame.grid(row=0, column=0, sticky="ew")
        
        # Title Section
        title_frame = tk.Frame(header_frame, bg="#0b0f19")
        title_frame.pack(side="left")
        
        lbl_main = tk.Label(title_frame, text="SAPID license analyser", font=("Helvetica", 20, "bold"), bg="#0b0f19", fg="#f8fafc")
        lbl_main.pack(anchor="w")
        lbl_sub = tk.Label(title_frame, text="Interactive SAP User Licensing & Logon Activity Dashboard", font=("Helvetica", 9), bg="#0b0f19", fg="#94a3b8")
        lbl_sub.pack(anchor="w")
        
        # File Action & Sync Box
        action_frame = tk.Frame(header_frame, bg="#0b0f19")
        action_frame.pack(side="right", fill="y")
        
        # URL Input & Sync
        sync_frame = tk.Frame(action_frame, bg="#1e293b", highlightbackground="#2a354f", highlightthickness=1)
        sync_frame.pack(side="left", padx=5, ipady=1)
        
        tk.Label(sync_frame, text="🔗", font=("Helvetica", 9), bg="#1e293b", fg="#94a3b8").pack(side="left", padx=(8, 2))
        
        self.ent_url = tk.Entry(sync_frame, font=("Helvetica", 9), bg="#1e293b", fg="#f8fafc", insertbackground="#f8fafc", bd=0, width=32)
        self.ent_url.pack(side="left", padx=5, ipady=4)
        self.ent_url.insert(0, self.active_url)
        self.ent_url.bind("<Return>", lambda e: self.trigger_sync())
        
        self.btn_sync = tk.Button(sync_frame, text="Sync", font=("Helvetica", 8, "bold"), bg="#0ea5e9", fg="#0b0f19", activebackground="#38bdf8", activeforeground="#0b0f19", bd=0, padx=12, pady=4, cursor="hand2", command=self.trigger_sync)
        self.btn_sync.pack(side="left")
        
        # Local upload button
        self.btn_upload = tk.Button(action_frame, text="📂 Upload Excel / CSV", font=("Helvetica", 8, "bold"), bg="#161e31", fg="#f8fafc", activebackground="#1f2945", activeforeground="#f8fafc", bd=1, highlightbackground="#2a354f", relief="solid", padx=12, pady=5, cursor="hand2", command=self.trigger_upload)
        self.btn_upload.pack(side="left", padx=8)
        
        # Clear/Reset Data
        self.btn_clear = tk.Button(action_frame, text="🗑️ Reset", font=("Helvetica", 8, "bold"), bg="#ef4444", fg="#f8fafc", activebackground="#dc2626", activeforeground="#f8fafc", bd=0, padx=12, pady=5, cursor="hand2", command=self.trigger_reset)
        self.btn_clear.pack(side="left")
        
        # ─── NAVIGATION ROW ───
        nav_frame = tk.Frame(self, bg="#0b0f19", padx=20, pady=5)
        nav_frame.grid(row=1, column=0, sticky="ew")
        
        self.btn_tab_dash = tk.Button(nav_frame, text="📊 Dashboard", font=("Helvetica", 10, "bold"), bg="#0ea5e9", fg="#0b0f19", bd=0, padx=18, pady=6, cursor="hand2", command=lambda: self.switch_tab("dashboard"))
        self.btn_tab_dash.pack(side="left")
        
        self.btn_tab_ledg = tk.Button(nav_frame, text="📋 Employee Ledger", font=("Helvetica", 10, "bold"), bg="#161e31", fg="#f8fafc", bd=0, padx=18, pady=6, cursor="hand2", command=lambda: self.switch_tab("ledger"))
        self.btn_tab_ledg.pack(side="left", padx=10)
        
        # Divider line
        div = tk.Frame(self, bg="#2a354f", height=1)
        div.grid(row=1, column=0, sticky="ews")
        
        # ─── CONTAINER FRAMES ───
        self.container = tk.Frame(self, bg="#0b0f19")
        self.container.grid(row=2, column=0, sticky="nsew", padx=20, pady=(15, 10))
        
        self.build_dashboard_page()
        self.build_ledger_page()
        
        # Show Dashboard initially
        self.switch_tab("dashboard")
        
        # ─── FOOTER ROW ───
        footer_frame = tk.Frame(self, bg="#0b0f19", padx=20, pady=12)
        footer_frame.grid(row=3, column=0, sticky="ew")
        
        # Footer Divider
        foot_div = tk.Frame(footer_frame, bg="#2a354f", height=1)
        foot_div.place(x=0, y=0, relwidth=1)
        
        # Left labels & modal links
        left_foot = tk.Frame(footer_frame, bg="#0b0f19")
        left_foot.pack(side="left", pady=(5, 0))
        
        tk.Label(left_foot, text="SAPID license analyser v1.0.0", font=("Helvetica", 8), bg="#0b0f19", fg="#64748b").pack(side="left")
        tk.Label(left_foot, text="|", font=("Helvetica", 8), bg="#0b0f19", fg="#2a354f").pack(side="left", padx=8)
        
        btn_man = tk.Button(left_foot, text="📖 User Manual", font=("Helvetica", 8), bg="#0b0f19", fg="#94a3b8", activebackground="#0b0f19", activeforeground="#f8fafc", bd=0, cursor="hand2", command=self.show_user_manual)
        btn_man.pack(side="left")
        tk.Label(left_foot, text="|", font=("Helvetica", 8), bg="#0b0f19", fg="#2a354f").pack(side="left", padx=8)
        
        btn_priv = tk.Button(left_foot, text="🔒 Privacy Policy", font=("Helvetica", 8), bg="#0b0f19", fg="#94a3b8", activebackground="#0b0f19", activeforeground="#f8fafc", bd=0, cursor="hand2", command=self.show_privacy_policy)
        btn_priv.pack(side="left")
        tk.Label(left_foot, text="|", font=("Helvetica", 8), bg="#0b0f19", fg="#2a354f").pack(side="left", padx=8)
        
        btn_lic = tk.Button(left_foot, text="📄 Licenses", font=("Helvetica", 8), bg="#0b0f19", fg="#94a3b8", activebackground="#0b0f19", activeforeground="#f8fafc", bd=0, cursor="hand2", command=self.show_licenses)
        btn_lic.pack(side="left")
        
        # Right action buttons
        right_foot = tk.Frame(footer_frame, bg="#0b0f19")
        right_foot.pack(side="right", pady=(5, 0))
        
        btn_export = tk.Button(right_foot, text="📥 Export Config", font=("Helvetica", 8, "bold"), bg="#1e293b", fg="#f8fafc", activebackground="#334155", activeforeground="#f8fafc", bd=0, padx=12, pady=5, cursor="hand2", command=self.export_config)
        btn_export.pack(side="left")
        
        btn_import = tk.Button(right_foot, text="📤 Import Config", font=("Helvetica", 8, "bold"), bg="#1e293b", fg="#f8fafc", activebackground="#334155", activeforeground="#f8fafc", bd=0, padx=12, pady=5, cursor="hand2", command=self.import_config)
        btn_import.pack(side="left", padx=8)
        
        btn_logs = tk.Button(right_foot, text="🪵 Error Logs", font=("Helvetica", 8, "bold"), bg="#161e31", fg="#f43f5e", activebackground="#ef4444", activeforeground="#0b0f19", bd=1, highlightbackground="#f43f5e", relief="solid", padx=12, pady=4, cursor="hand2", command=self.show_logs_modal)
        btn_logs.pack(side="left")

    def build_dashboard_page(self):
        self.frame_dash = tk.Frame(self.container, bg="#0b0f19")
        
        # Main dashboard grid layout
        self.frame_dash.grid_rowconfigure(0, weight=0)  # KPIs row
        self.frame_dash.grid_rowconfigure(1, weight=0)  # Filter bar row
        self.frame_dash.grid_rowconfigure(2, weight=0)  # Charts row
        self.frame_dash.grid_rowconfigure(3, weight=1)  # Explorer & progress split row
        self.frame_dash.grid_columnconfigure(0, weight=1)
        
        # ─── 1. KPIs Row ───
        kpi_row = tk.Frame(self.frame_dash, bg="#0b0f19")
        kpi_row.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        for col_idx in range(4):
            kpi_row.grid_columnconfigure(col_idx, weight=1, uniform="kpi")
            
        self.kpi_total_users = KPICard(kpi_row, "Total Users", "0", "📊", "#0ea5e9")
        self.kpi_total_users.grid(row=0, column=0, sticky="ew", padx=(0, 8))
        
        self.kpi_total_purchased = KPICard(kpi_row, "Total Purchased", "0", "👥", "#6366f1")
        self.kpi_total_purchased.grid(row=0, column=1, sticky="ew", padx=8)
        
        self.kpi_available = KPICard(kpi_row, "Available Licenses", "0", "📈", "#a855f7")
        self.kpi_available.grid(row=0, column=2, sticky="ew", padx=8)
        
        self.kpi_top_license = KPICard(kpi_row, "Top License Type", "-", "🏢", "#10b981")
        self.kpi_top_license.grid(row=0, column=3, sticky="ew", padx=(8, 0))
        
        # ─── 2. Filter Bar (inline panel matching web card) ───
        filter_panel = tk.Frame(self.frame_dash, bg="#161e31", highlightbackground="#2a354f", highlightthickness=1)
        filter_panel.grid(row=1, column=0, sticky="ew", pady=(0, 10))  # Placed in row 1, no overlap
        
        lbl_f_lic = tk.Label(filter_panel, text="LICENSE ID", font=("Helvetica", 7, "bold"), bg="#161e31", fg="#64748b")
        lbl_f_lic.grid(row=0, column=0, sticky="w", padx=(15, 5), pady=(8, 0))
        self.dash_filter_lic_var = tk.StringVar(value="ALL")
        self.dash_opt_lic = ttk.OptionMenu(filter_panel, self.dash_filter_lic_var, "ALL", command=lambda v: self.apply_dash_filters())
        self.dash_opt_lic.config(width=16)
        self.dash_opt_lic.grid(row=1, column=0, sticky="w", padx=(15, 5), pady=(0, 10))
        
        lbl_f_dep = tk.Label(filter_panel, text="DEPARTMENT", font=("Helvetica", 7, "bold"), bg="#161e31", fg="#64748b")
        lbl_f_dep.grid(row=0, column=1, sticky="w", padx=(10, 5), pady=(8, 0))
        self.dash_filter_dept_var = tk.StringVar(value="ALL")
        self.dash_opt_dept = ttk.OptionMenu(filter_panel, self.dash_filter_dept_var, "ALL", command=lambda v: self.apply_dash_filters())
        self.dash_opt_dept.config(width=18)
        self.dash_opt_dept.grid(row=1, column=1, sticky="w", padx=(10, 5), pady=(0, 10))
        
        # ─── 3. Visualization Canvas row ───
        charts_row = tk.Frame(self.frame_dash, bg="#0b0f19")
        charts_row.grid(row=2, column=0, sticky="ew", pady=(0, 15))  # Placed in row 2
        for col_idx in range(3):
            charts_row.grid_columnconfigure(col_idx, weight=1, uniform="chart")
            
        # Chart 1 Card: License Distribution
        c1_card = tk.Frame(charts_row, bg="#161e31", highlightbackground="#2a354f", highlightthickness=1)
        c1_card.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        tk.Label(c1_card, text="License Distribution", font=("Helvetica", 10, "bold"), bg="#161e31", fg="#f8fafc", anchor="w").pack(fill="x", padx=15, pady=(12, 1))
        tk.Label(c1_card, text="Total assigned users by License ID", font=("Helvetica", 7), bg="#161e31", fg="#94a3b8", anchor="w").pack(fill="x", padx=15, pady=(0, 8))
        self.can_lic = tk.Canvas(c1_card, bg="#161e31", bd=0, highlightthickness=0, height=220)
        self.can_lic.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        self.tooltip_lic = CanvasTooltip(self.can_lic)
        self.can_lic.bind("<Motion>", lambda e: self.on_chart_motion(e, self.can_lic, self.lic_bars_map, self.tooltip_lic))
        self.lic_bars_map = {}
        
        # Chart 2 Card: Dept Distribution
        c2_card = tk.Frame(charts_row, bg="#161e31", highlightbackground="#2a354f", highlightthickness=1)
        c2_card.grid(row=0, column=1, sticky="nsew", padx=8)
        tk.Label(c2_card, text="Department Distribution", font=("Helvetica", 10, "bold"), bg="#161e31", fg="#f8fafc", anchor="w").pack(fill="x", padx=15, pady=(12, 1))
        tk.Label(c2_card, text="Active licenses accumulated by departments", font=("Helvetica", 7), bg="#161e31", fg="#94a3b8", anchor="w").pack(fill="x", padx=15, pady=(0, 8))
        self.can_dept = tk.Canvas(c2_card, bg="#161e31", bd=0, highlightthickness=0, height=220)
        self.can_dept.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        self.tooltip_dept = CanvasTooltip(self.can_dept)
        self.can_dept.bind("<Motion>", lambda e: self.on_chart_motion(e, self.can_dept, self.dept_bars_map, self.tooltip_dept))
        self.dept_bars_map = {}
        
        # Chart 3 Card: Plant ID Distribution
        c3_card = tk.Frame(charts_row, bg="#161e31", highlightbackground="#2a354f", highlightthickness=1)
        c3_card.grid(row=0, column=2, sticky="nsew", padx=(8, 0))
        tk.Label(c3_card, text="Plant ID Distribution", font=("Helvetica", 10, "bold"), bg="#161e31", fg="#f8fafc", anchor="w").pack(fill="x", padx=15, pady=(12, 1))
        tk.Label(c3_card, text="Users grouped by Plant ID (Click bars to drilldown)", font=("Helvetica", 7), bg="#161e31", fg="#94a3b8", anchor="w").pack(fill="x", padx=15, pady=(0, 8))
        self.can_plant = tk.Canvas(c3_card, bg="#161e31", bd=0, highlightthickness=0, height=220)
        self.can_plant.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        self.tooltip_plant = CanvasTooltip(self.can_plant)
        self.can_plant.bind("<Motion>", lambda e: self.on_chart_motion(e, self.can_plant, self.plant_bars_map, self.tooltip_plant))
        self.can_plant.bind("<Button-1>", self.on_plant_chart_click)
        self.plant_bars_map = {}
        
        # ─── 4. Explorer & License limits split row ───
        split_row = tk.Frame(self.frame_dash, bg="#0b0f19")
        split_row.grid(row=3, column=0, sticky="nsew")  # Placed in row 3
        split_row.grid_columnconfigure(0, weight=2)  # Group Explorer gets 65% width
        split_row.grid_columnconfigure(1, weight=1)  # Quota progress gets 35% width
        split_row.grid_rowconfigure(0, weight=1)
        
        # Left Panel: SAP User Group Explorer
        exp_card = tk.Frame(split_row, bg="#161e31", highlightbackground="#2a354f", highlightthickness=1)
        exp_card.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        
        exp_header = tk.Frame(exp_card, bg="#161e31")
        exp_header.pack(fill="x", padx=15, pady=(12, 8))
        
        header_text_frame = tk.Frame(exp_header, bg="#161e31")
        header_text_frame.pack(side="left")
        tk.Label(header_text_frame, text="SAP User Group Explorer", font=("Helvetica", 11, "bold"), bg="#161e31", fg="#f8fafc", anchor="w").pack(fill="x")
        tk.Label(header_text_frame, text="Select a group below to display active members", font=("Helvetica", 7), bg="#161e31", fg="#94a3b8", anchor="w").pack(fill="x")
        
        # Explorer License Filter
        exp_filter_frame = tk.Frame(exp_header, bg="#161e31")
        exp_filter_frame.pack(side="right")
        tk.Label(exp_filter_frame, text="LICENSE: ", font=("Helvetica", 7, "bold"), bg="#161e31", fg="#64748b").pack(side="left")
        self.exp_filter_lic_var = tk.StringVar(value="ALL")
        self.exp_opt_lic = ttk.OptionMenu(exp_filter_frame, self.exp_filter_lic_var, "ALL", command=lambda v: self.refresh_explorer())
        self.exp_opt_lic.config(width=10)
        self.exp_opt_lic.pack(side="left")
        
        # Split pane inside Group Explorer
        exp_split = tk.Frame(exp_card, bg="#161e31")
        exp_split.pack(fill="both", expand=True, padx=15, pady=(0, 12))
        exp_split.grid_columnconfigure(0, weight=1)
        exp_split.grid_columnconfigure(1, weight=1)
        exp_split.grid_rowconfigure(0, weight=1)
        
        # Left Side: Functional Group list frame
        self.exp_btn_container = tk.Frame(exp_split, bg="#161e31")
        self.exp_btn_container.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        
        # Right Side: Assigned Users Table
        users_table_card = tk.Frame(exp_split, bg="#0f172a", bd=1, relief="solid", highlightthickness=0)
        users_table_card.grid(row=0, column=1, sticky="nsew")
        
        self.exp_lbl_table_title = tk.Label(users_table_card, text="Assigned Users", font=("Helvetica", 9, "bold"), bg="#0f172a", fg="#f8fafc", anchor="w")
        self.exp_lbl_table_title.pack(fill="x", padx=10, pady=(10, 2))
        self.exp_lbl_table_sub = tk.Label(users_table_card, text="Select a group to load members", font=("Helvetica", 7), bg="#0f172a", fg="#94a3b8", anchor="w")
        self.exp_lbl_table_sub.pack(fill="x", padx=10, pady=(0, 8))
        
        exp_tbl_frame = tk.Frame(users_table_card, bg="#0f172a")
        exp_tbl_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        exp_cols = ("name", "license", "department")
        self.exp_tree = ttk.Treeview(exp_tbl_frame, columns=exp_cols, show="headings", height=8, selectmode="browse")
        self.exp_tree.heading("name", text="User Name")
        self.exp_tree.heading("license", text="License")
        self.exp_tree.heading("department", text="Department")
        self.exp_tree.column("name", width=120, anchor="w")
        self.exp_tree.column("license", width=70, anchor="center")
        self.exp_tree.column("department", width=90, anchor="center")
        self.exp_tree.pack(side="left", fill="both", expand=True)
        
        exp_scroll = ttk.Scrollbar(exp_tbl_frame, orient="vertical", command=self.exp_tree.yview)
        exp_scroll.pack(side="right", fill="y")
        self.exp_tree.config(yscrollcommand=exp_scroll.set)
        
        self.exp_tree.bind("<Double-1>", self.on_explorer_user_double_click)
        
        # Right Panel: License Limits sidebar
        quota_card = tk.Frame(split_row, bg="#161e31", highlightbackground="#2a354f", highlightthickness=1)
        quota_card.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
        
        tk.Label(quota_card, text="License Distribution Progress", font=("Helvetica", 11, "bold"), bg="#161e31", fg="#f8fafc", anchor="w").pack(fill="x", padx=15, pady=(12, 1))
        tk.Label(quota_card, text="Remaining count and limits edit", font=("Helvetica", 7), bg="#161e31", fg="#94a3b8", anchor="w").pack(fill="x", padx=15, pady=(0, 10))
        
        # Scrollable container for limit widgets
        sidebar_scroll_frame = tk.Frame(quota_card, bg="#161e31")
        sidebar_scroll_frame.pack(fill="both", expand=True, padx=15, pady=(0, 10))
        
        # Scrollable canvas setup
        self.sidebar_canvas = tk.Canvas(sidebar_scroll_frame, bg="#161e31", bd=0, highlightthickness=0)
        self.sidebar_scrollbar = ttk.Scrollbar(sidebar_scroll_frame, orient="vertical", command=self.sidebar_canvas.yview)
        self.scrollable_sidebar = tk.Frame(self.sidebar_canvas, bg="#161e31")
        
        self.scrollable_sidebar.bind(
            "<Configure>",
            lambda e: self.sidebar_canvas.configure(
                scrollregion=self.sidebar_canvas.bbox("all")
            )
        )
        self.sidebar_canvas_window = self.sidebar_canvas.create_window((0, 0), window=self.scrollable_sidebar, anchor="nw")
        
        # Stretch inner frame to match Canvas width on Canvas resize
        self.sidebar_canvas.bind(
            "<Configure>",
            lambda e: self.sidebar_canvas.itemconfig(
                self.sidebar_canvas_window,
                width=e.width
            )
        )
        self.sidebar_canvas.configure(yscrollcommand=self.sidebar_scrollbar.set)
        
        # Bind mouse wheel to canvas for scrolling
        self.sidebar_canvas.bind("<Enter>", self._bind_sidebar_mousewheel)
        self.sidebar_canvas.bind("<Leave>", self._unbind_sidebar_mousewheel)
        
        self.sidebar_canvas.pack(side="left", fill="both", expand=True)
        self.sidebar_scrollbar.pack(side="right", fill="y")
        
        # Place progress widgets
        self.build_quota_sidebar()

    def _bind_sidebar_mousewheel(self, event):
        self.sidebar_canvas.bind_all("<MouseWheel>", self._on_sidebar_mousewheel)
        self.sidebar_canvas.bind_all("<Button-4>", self._on_sidebar_mousewheel_linux)
        self.sidebar_canvas.bind_all("<Button-5>", self._on_sidebar_mousewheel_linux)
        
    def _unbind_sidebar_mousewheel(self, event):
        self.sidebar_canvas.unbind_all("<MouseWheel>")
        self.sidebar_canvas.unbind_all("<Button-4>")
        self.sidebar_canvas.unbind_all("<Button-5>")
        
    def _on_sidebar_mousewheel(self, event):
        self.sidebar_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        
    def _on_sidebar_mousewheel_linux(self, event):
        if event.num == 4:
            self.sidebar_canvas.yview_scroll(-1, "units")
        elif event.num == 5:
            self.sidebar_canvas.yview_scroll(1, "units")

    def build_quota_sidebar(self):
        # Empty previous
        for w in self.scrollable_sidebar.winfo_children():
            w.destroy()
            
        self.quota_widgets = {}
        licenses = ["AX", "AY", "FX", "HC", "HD", "Other"]
        
        for idx, lic in enumerate(licenses):
            frame = tk.Frame(self.scrollable_sidebar, bg="#161e31", pady=6)
            frame.pack(fill="x", expand=True)
            
            # Text headers
            lbl_lic = tk.Label(frame, text=lic, font=("Helvetica", 9, "bold"), bg="#161e31", fg=LICENSE_COLORS.get(lic, "#fff"), anchor="w")
            lbl_lic.pack(fill="x")
            
            # Details subrow (allocated / limits input)
            sub_row = tk.Frame(frame, bg="#161e31")
            sub_row.pack(fill="x", pady=2)
            
            lbl_detail = tk.Label(sub_row, text="0 / 0 Assigned (0%)", font=("Helvetica", 8), bg="#161e31", fg="#94a3b8", anchor="w")
            lbl_detail.pack(side="left")
            
            # Input limit entry
            limit_entry_frame = tk.Frame(sub_row, bg="#1e293b", highlightbackground="#2a354f", highlightthickness=1)
            limit_entry_frame.pack(side="right")
            
            tk.Label(limit_entry_frame, text="Limit:", font=("Helvetica", 7), bg="#1e293b", fg="#64748b").pack(side="left", padx=(5, 2))
            ent_lim = tk.Entry(limit_entry_frame, font=("Helvetica", 8), bg="#1e293b", fg="#f8fafc", bd=0, width=5, justify="center")
            ent_lim.pack(side="left", ipady=2, padx=(0, 2))
            # Bind Entry edit
            ent_lim.bind("<FocusOut>", lambda e, l=lic: self.update_license_limit(l))
            ent_lim.bind("<Return>", lambda e, l=lic: self.update_license_limit(l))
            
            # Remaining line
            rem_row = tk.Frame(frame, bg="#161e31")
            rem_row.pack(fill="x")
            tk.Label(rem_row, text="Remaining Count:", font=("Helvetica", 7), bg="#161e31", fg="#64748b").pack(side="left")
            lbl_rem = tk.Label(rem_row, text="0", font=("Helvetica", 8, "bold"), bg="#161e31", fg="#10b981")
            lbl_rem.pack(side="right")
            
            # Canvas progress line bar
            bar_can = tk.Canvas(frame, bg="#1e293b", height=8, bd=0, highlightthickness=0)
            bar_can.pack(fill="x", pady=(4, 0))
            
            self.quota_widgets[lic] = {
                "detail_lbl": lbl_detail,
                "entry": ent_lim,
                "rem_lbl": lbl_rem,
                "canvas": bar_can
            }

    def build_ledger_page(self):
        self.frame_ledg = tk.Frame(self.container, bg="#0b0f19")
        
        self.frame_ledg.grid_rowconfigure(0, weight=0)  # Search/Filter pane
        self.frame_ledg.grid_rowconfigure(1, weight=1)  # Table view pane
        self.frame_ledg.grid_rowconfigure(2, weight=0)  # Pagination pane
        self.frame_ledg.grid_columnconfigure(0, weight=1)
        
        # ─── 1. Search & Filter Bar ───
        filt_panel = tk.Frame(self.frame_ledg, bg="#161e31", highlightbackground="#2a354f", highlightthickness=1, padx=15, pady=10)
        filt_panel.grid(row=0, column=0, sticky="ew", pady=(0, 15))
        
        # Search Box
        lbl_s = tk.Label(filt_panel, text="SEARCH USER", font=("Helvetica", 7, "bold"), bg="#161e31", fg="#64748b")
        lbl_s.grid(row=0, column=0, sticky="w", padx=(0, 5), pady=(2, 0))
        
        search_input_frame = tk.Frame(filt_panel, bg="#1e293b", highlightbackground="#2a354f", highlightthickness=1)
        search_input_frame.grid(row=1, column=0, sticky="w", padx=(0, 5), pady=(0, 5))
        
        tk.Label(search_input_frame, text="🔍", font=("Helvetica", 8), bg="#1e293b", fg="#64748b").pack(side="left", padx=(8, 2))
        self.ent_search = tk.Entry(search_input_frame, font=("Helvetica", 9), bg="#1e293b", fg="#f8fafc", insertbackground="#f8fafc", bd=0, width=22)
        self.ent_search.pack(side="left", ipady=3, padx=(0, 5))
        self.ent_search.bind("<KeyRelease>", lambda e: self.trigger_search_ledger())
        
        # License OptionMenu
        lbl_l = tk.Label(filt_panel, text="LICENSE ID", font=("Helvetica", 7, "bold"), bg="#161e31", fg="#64748b")
        lbl_l.grid(row=0, column=1, sticky="w", padx=10, pady=(2, 0))
        
        self.ledg_filter_lic_var = tk.StringVar(value="ALL")
        self.ledg_opt_lic = ttk.OptionMenu(filt_panel, self.ledg_filter_lic_var, "ALL", command=lambda v: self.apply_ledger_filters())
        self.ledg_opt_lic.config(width=12)
        self.ledg_opt_lic.grid(row=1, column=1, sticky="w", padx=10, pady=(0, 5))
        
        # Dept OptionMenu
        lbl_d = tk.Label(filt_panel, text="DEPARTMENT", font=("Helvetica", 7, "bold"), bg="#161e31", fg="#64748b")
        lbl_d.grid(row=0, column=2, sticky="w", padx=10, pady=(2, 0))
        
        self.ledg_filter_dept_var = tk.StringVar(value="ALL")
        self.ledg_opt_dept = ttk.OptionMenu(filt_panel, self.ledg_filter_dept_var, "ALL", command=lambda v: self.apply_ledger_filters())
        self.ledg_opt_dept.config(width=15)
        self.ledg_opt_dept.grid(row=1, column=2, sticky="w", padx=10, pady=(0, 5))
        
        # Clear Filter Button
        self.btn_reset_filters = tk.Button(filt_panel, text="🔄 Clear", font=("Helvetica", 8, "bold"), bg="#334155", fg="#f8fafc", activebackground="#475569", activeforeground="#f8fafc", bd=0, padx=12, pady=5, cursor="hand2", command=self.reset_ledger_filters)
        self.btn_reset_filters.grid(row=1, column=3, sticky="w", padx=10, pady=(0, 5))
        
        # Rows Page Limit OptionMenu on Right
        lbl_lim = tk.Label(filt_panel, text="ROWS PER PAGE", font=("Helvetica", 7, "bold"), bg="#161e31", fg="#64748b")
        lbl_lim.grid(row=0, column=4, sticky="e", padx=(0, 10), pady=(2, 0))
        filt_panel.grid_columnconfigure(4, weight=1)
        
        self.ledg_limit_var = tk.StringVar(value="10")
        self.ledg_opt_limit = ttk.OptionMenu(filt_panel, self.ledg_limit_var, "10", "5", "10", "25", "50", command=lambda v: self.apply_ledger_page_size())
        self.ledg_opt_limit.config(width=6)
        self.ledg_opt_limit.grid(row=1, column=4, sticky="e", padx=(0, 10), pady=(0, 5))
        
        # ─── 2. Ledger Treeview Table ───
        table_card = tk.Frame(self.frame_ledg, bg="#161e31", highlightbackground="#2a354f", highlightthickness=1)
        table_card.grid(row=1, column=0, sticky="nsew", pady=(0, 10))
        
        table_container = tk.Frame(table_card, bg="#161e31")
        table_container.pack(fill="both", expand=True, padx=15, pady=15)
        
        cols = ("sapid", "name", "license", "department", "group", "lastlogon")
        self.tree_ledg = ttk.Treeview(table_container, columns=cols, show="headings", selectmode="browse")
        
        # Header text bindings for click-to-sort columns
        self.tree_ledg.heading("sapid", text="SAPID", command=lambda: self.sort_ledger("sapid"))
        self.tree_ledg.heading("name", text="User Name", command=lambda: self.sort_ledger("name"))
        self.tree_ledg.heading("license", text="License", command=lambda: self.sort_ledger("license"))
        self.tree_ledg.heading("department", text="Department", command=lambda: self.sort_ledger("department"))
        self.tree_ledg.heading("group", text="User Group", command=lambda: self.sort_ledger("group"))
        self.tree_ledg.heading("lastlogon", text="Last Logon", command=lambda: self.sort_ledger("lastlogon"))
        
        self.tree_ledg.column("sapid", width=120, anchor="center")
        self.tree_ledg.column("name", width=220, anchor="w")
        self.tree_ledg.column("license", width=110, anchor="center")
        self.tree_ledg.column("department", width=140, anchor="center")
        self.tree_ledg.column("group", width=180, anchor="w")
        self.tree_ledg.column("lastlogon", width=150, anchor="center")
        
        self.tree_ledg.pack(side="left", fill="both", expand=True)
        
        # Scrollbars
        ysb = ttk.Scrollbar(table_container, orient="vertical", command=self.tree_ledg.yview)
        ysb.pack(side="right", fill="y")
        self.tree_ledg.config(yscrollcommand=ysb.set)
        
        self.tree_ledg.bind("<Double-1>", self.on_ledger_row_double_click)
        
        # ─── 3. Pagination Controls ───
        pag_frame = tk.Frame(self.frame_ledg, bg="#0b0f19")
        pag_frame.grid(row=2, column=0, sticky="ew")
        
        self.lbl_pag_info = tk.Label(pag_frame, text="Showing 0 to 0 of 0 employees", font=("Helvetica", 9), bg="#0b0f19", fg="#94a3b8")
        self.lbl_pag_info.pack(side="left")
        
        pag_buttons_frame = tk.Frame(pag_frame, bg="#0b0f19")
        pag_buttons_frame.pack(side="right")
        
        self.btn_pag_prev = tk.Button(pag_buttons_frame, text="Previous", font=("Helvetica", 8, "bold"), bg="#1e293b", fg="#f8fafc", activebackground="#334155", activeforeground="#f8fafc", bd=0, padx=12, pady=5, cursor="hand2", command=self.ledger_prev_page)
        self.btn_pag_prev.pack(side="left", padx=(0, 5))
        
        self.btn_pag_next = tk.Button(pag_buttons_frame, text="Next", font=("Helvetica", 8, "bold"), bg="#1e293b", fg="#f8fafc", activebackground="#334155", activeforeground="#f8fafc", bd=0, padx=12, pady=5, cursor="hand2", command=self.ledger_next_page)
        self.btn_pag_next.pack(side="left")

    # ─── UI Actions & Tab Switching ───────────────────────────────────────────
    def switch_tab(self, tab_name):
        self.current_tab = tab_name
        if tab_name == "dashboard":
            self.frame_ledg.grid_forget()
            self.frame_dash.grid(row=0, column=0, sticky="nsew")
            
            # Nav buttons coloring
            self.btn_tab_dash.config(bg="#0ea5e9", fg="#0b0f19")
            self.btn_tab_ledg.config(bg="#161e31", fg="#f8fafc")
            self.refresh_dashboard_view()
        else:
            self.frame_dash.grid_forget()
            self.frame_ledg.grid(row=0, column=0, sticky="nsew")
            
            self.btn_tab_dash.config(bg="#161e31", fg="#f8fafc")
            self.btn_tab_ledg.config(bg="#0ea5e9", fg="#0b0f19")
            self.refresh_ledger_view()

    def load_config(self):
        # Default config state
        self.active_url = ""
        self.raw_data = []
        self.limits = DEFAULT_LIMITS.copy()
        
        if os.path.exists(CONFIG_PATH):
            try:
                with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                    config = json.load(f)
                self.active_url = config.get("dashboard_sheet_url", "")
                self.raw_data = config.get("cached_sheet_data", [])
                self.limits = config.get("limits", DEFAULT_LIMITS.copy())
            except Exception as e:
                append_log("ERROR", f"Config load failed: {e}")
                
        # Fill mock data if cache empty
        if not self.raw_data:
            self.raw_data = generate_mock_data()
            self.save_config()

    def save_config(self):
        try:
            config = {
                "dashboard_sheet_url": self.active_url,
                "cached_sheet_data": self.raw_data,
                "limits": self.limits
            }
            with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(config, f, indent=2)
        except Exception as e:
            append_log("ERROR", f"Config save failed: {e}")

    # ─── Data Syncing ─────────────────────────────────────────────────────────
    def trigger_sync(self):
        url = self.ent_url.get().strip()
        if not url:
            messagebox.showwarning("Sync Input", "Please provide a valid spreadsheet URL.")
            return
        
        self.btn_sync.config(state="disabled", text="Syncing...")
        self.btn_upload.config(state="disabled")
        self.btn_clear.config(state="disabled")
        
        # Fetch data in background thread
        threading.Thread(target=self.bg_fetch_sheet, args=(url,), daemon=True).start()
        
    def bg_fetch_sheet(self, url):
        # Extract spreadsheet CSV or Apps Script target
        is_apps_script = "script.google.com" in url
        target_url = url
        if not is_apps_script:
            target_url = get_sheets_csv_url(url)
            if not target_url:
                self.after(0, lambda: self.on_fetch_failed("Could not extract Sheet ID from link format."))
                return
                
        try:
            # Standard request
            req = urllib.request.Request(
                target_url,
                headers={"User-Agent": "SAPID-license-analyser-Desktop/1.0"}
            )
            with urllib.request.urlopen(req, timeout=15) as res:
                body_bytes = res.read()
                
            if is_apps_script:
                # Expect JSON payload
                json_data = json.loads(body_bytes.decode("utf-8"))
                parsed_rows = []
                # Check for 2D array representation
                if isinstance(json_data, list) and len(json_data) > 0 and isinstance(json_data[0], list):
                    headers = json_data[0]
                    for r in json_data[1:]:
                        row_dict = {}
                        for col_idx, header in enumerate(headers):
                            row_dict[header] = r[col_idx] if col_idx < len(r) else ""
                        parsed_rows.append(row_dict)
                elif isinstance(json_data, list):
                    parsed_rows = json_data
                else:
                    raise Exception("Unexpected script response format.")
                self.after(0, lambda: self.on_fetch_success(url, parsed_rows))
            else:
                # Expect CSV text
                csv_text = body_bytes.decode("utf-8")
                f = csv.DictReader(csv_text.splitlines())
                parsed_rows = list(f)
                if not parsed_rows:
                    raise Exception("Parsed spreadsheet CSV data is empty.")
                self.after(0, lambda: self.on_fetch_success(url, parsed_rows))
                
        except Exception as e:
            self.after(0, lambda: self.on_fetch_failed(str(e)))
            
    def on_fetch_success(self, original_url, rows):
        self.btn_sync.config(state="normal", text="Sync")
        self.btn_upload.config(state="normal")
        self.btn_clear.config(state="normal")
        
        self.active_url = original_url
        self.raw_data = rows
        self.save_config()
        self.refresh_ui()
        self.show_toast("Sync Completed Successfully!")
        
    def on_fetch_failed(self, err_msg):
        self.btn_sync.config(state="normal", text="Sync")
        self.btn_upload.config(state="normal")
        self.btn_clear.config(state="normal")
        
        append_log("ERROR", f"Network sync failed: {err_msg}")
        messagebox.showerror("Sync Failed", f"Spreadsheet fetch error:\n{err_msg}")

    # ─── Local Excel / CSV Uploader ───────────────────────────────────────────
    def trigger_upload(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel & CSV Sheets", "*.csv *.xlsx *.xls"), ("CSV File", "*.csv"), ("Excel File", "*.xlsx *.xls")]
        )
        if not file_path:
            return
            
        try:
            if file_path.lower().endswith(".csv"):
                with open(file_path, "r", encoding="utf-8") as f:
                    reader = csv.DictReader(f)
                    parsed_rows = list(reader)
                if not parsed_rows:
                    raise Exception("CSV file content is empty.")
                self.on_upload_success(parsed_rows)
            else:
                # Excel file
                if not HAS_OPENPYXL:
                    messagebox.showerror(
                        "Package Missing",
                        "Excel parser requires package 'openpyxl'.\n"
                        "Please run in terminal:\n\npip install openpyxl\n\nOr import file as standard .CSV format."
                    )
                    return
                wb = openpyxl.load_workbook(file_path, data_only=True)
                # Find sheet based on target priority keywords
                sheet_name = wb.sheetnames[0]
                targets = ['live users', 'live user', 'acctive user', 'active user', 'sapidinfo', 'sapid info', 'sapid', 'total id']
                for t in targets:
                    matched = next((name for name in wb.sheetnames if name.strip().lower() == t), None)
                    if matched:
                        sheet_name = matched
                        break
                sheet = wb[sheet_name]
                # Extract columns
                rows_iter = sheet.iter_rows(values_only=True)
                headers = next(rows_iter, None)
                if not headers:
                    raise Exception("Excel sheet header row is empty.")
                parsed_rows = []
                for r in rows_iter:
                    if not any(cell is not None for cell in r):
                        continue  # Skip blank row
                    row_dict = {}
                    for idx, header in enumerate(headers):
                        if header is not None:
                            row_dict[header] = r[idx] if idx < len(r) else ""
                    parsed_rows.append(row_dict)
                self.on_upload_success(parsed_rows)
        except Exception as e:
            append_log("ERROR", f"File parse failed: {e}")
            messagebox.showerror("Parse Error", f"Failed to load spreadsheet:\n{e}")

    def on_upload_success(self, rows):
        self.raw_data = rows
        # Keep URL, but clear sheet link indicator since it is loaded from client file
        self.active_url = ""
        self.ent_url.delete(0, "end")
        self.save_config()
        self.refresh_ui()
        self.show_toast("Local spreadsheet imported successfully!")

    def trigger_reset(self):
        if messagebox.askyesno("Reset Data", "Restore preloaded SAP user dataset?"):
            self.active_url = ""
            self.ent_url.delete(0, "end")
            self.raw_data = generate_mock_data()
            self.save_config()
            self.refresh_ui()
            self.show_toast("Data reset to default dataset.")

    # ─── Filter Calculations ──────────────────────────────────────────────────
    def apply_dash_filters(self):
        self.refresh_dashboard_view()
        
    def apply_ledger_filters(self):
        self.ledger_page = 1
        self.refresh_ledger_view()
        
    def apply_ledger_page_size(self):
        size = self.ledg_limit_var.get()
        try:
            self.ledger_page_size = int(size)
        except ValueError:
            self.ledger_page_size = 10
        self.ledger_page = 1
        self.refresh_ledger_view()

    def reset_ledger_filters(self):
        self.ent_search.delete(0, "end")
        self.ledg_filter_lic_var.set("ALL")
        self.ledg_filter_dept_var.set("ALL")
        self.ledger_page = 1
        self.refresh_ledger_view()

    def trigger_search_ledger(self):
        self.ledger_page = 1
        self.refresh_ledger_view()

    def update_license_limit(self, lic):
        widgets = self.quota_widgets.get(lic)
        if not widgets:
            return
        entry = widgets["entry"]
        val_str = entry.get().strip()
        try:
            val = int(val_str)
            if val < 0:
                raise ValueError
        except ValueError:
            val = DEFAULT_LIMITS.get(lic, 150)
            entry.delete(0, "end")
            entry.insert(0, str(val))
            
        self.limits[lic] = val
        self.save_config()
        self.refresh_ui()
        self.show_toast(f"Limit for {lic} adjusted to {val}")

    # ─── Master Refresh ───────────────────────────────────────────────────────
    def refresh_ui(self):
        # Extract unique options lists for filter menus
        unique_licenses = sorted(list(set(get_row_value(r, LICENSE_KEYS, "Unknown") for r in self.raw_data)))
        unique_depts = sorted(list(set(get_row_value(r, DEPT_KEYS, "Unknown") for r in self.raw_data)))
        
        # Standard filter additions
        self.update_option_menu(self.dash_opt_lic, self.dash_filter_lic_var, ["ALL"] + unique_licenses)
        self.update_option_menu(self.dash_opt_dept, self.dash_filter_dept_var, ["ALL"] + unique_depts)
        self.update_option_menu(self.exp_opt_lic, self.exp_filter_lic_var, ["ALL"] + unique_licenses)
        
        self.update_option_menu(self.ledg_opt_lic, self.ledg_filter_lic_var, ["ALL"] + unique_licenses)
        self.update_option_menu(self.ledg_opt_dept, self.ledg_filter_dept_var, ["ALL"] + unique_depts)
        
        if self.current_tab == "dashboard":
            self.refresh_dashboard_view()
        else:
            self.refresh_ledger_view()

    def update_option_menu(self, menu_widget, str_var, options):
        # Safe OptionMenu update method
        menu = menu_widget["menu"]
        menu.delete(0, "end")
        for opt in options:
            menu.add_command(label=opt, command=lambda value=opt: str_var.set(value))
        # Keep selected option if valid
        if str_var.get() not in options:
            str_var.set(options[0])

    # ─── Dashboard Tab Rendering ──────────────────────────────────────────────
    def refresh_dashboard_view(self):
        # Filter raw rows
        lic_filter = self.dash_filter_lic_var.get()
        dept_filter = self.dash_filter_dept_var.get()
        
        self.filtered_data = []
        for r in self.raw_data:
            match_lic = (lic_filter == "ALL" or get_row_value(r, LICENSE_KEYS) == lic_filter)
            match_dept = (dept_filter == "ALL" or get_row_value(r, DEPT_KEYS) == dept_filter)
            if match_lic and match_dept:
                self.filtered_data.append(r)
                
        # Total Allocated Users
        total_allocated = len(self.filtered_data)
        
        # Calculate limits and remaining
        total_purchased = 0
        allocated_counts = {l: 0 for l in ["AX", "AY", "FX", "HC", "HD", "Other"]}
        for r in self.filtered_data:
            lic = get_row_value(r, LICENSE_KEYS, "Other")
            if lic not in allocated_counts:
                lic = "Other"
            allocated_counts[lic] += 1
            
        for l in ["AX", "AY", "FX", "HC", "HD", "Other"]:
            total_purchased += self.limits.get(l, DEFAULT_LIMITS[l])
            
        # Available
        available = max(0, total_purchased - len(self.raw_data)) # Total quota remaining
        
        # Top License
        top_license = "-"
        max_count = 0
        for l, cnt in allocated_counts.items():
            if cnt > max_count:
                max_count = cnt
                top_license = l
        top_lic_str = f"{top_license} ({max_count} users)" if max_count > 0 else "-"
        
        # Update KPI UI
        self.kpi_total_users.update_val(total_allocated)
        self.kpi_total_purchased.update_val(total_purchased)
        self.kpi_available.update_val(available)
        self.kpi_top_license.update_val(top_lic_str)
        
        self.allocated_counts = allocated_counts
        # Re-render Quota Sidebar widgets
        self.render_quota_sidebar(self.allocated_counts)
        
        # Render custom Canvas Charts
        self.render_charts()
        
        # Refresh SAP Explorer
        self.refresh_explorer()

    def render_quota_sidebar(self, allocated_counts):
        for lic, widgets in self.quota_widgets.items():
            count = allocated_counts.get(lic, 0)
            limit = self.limits.get(lic, DEFAULT_LIMITS.get(lic, 150))
            percentage = min(100, round((count / limit) * 100)) if limit > 0 else 0
            
            widgets["detail_lbl"].config(
                text=f"{count} / {limit} Assigned ({percentage}%)"
            )
            
            entry = widgets["entry"]
            if self.focus_get() != entry:
                entry.delete(0, "end")
                entry.insert(0, str(limit))
                
            rem = limit - count
            widgets["rem_lbl"].config(text=str(rem))
            if rem < 0:
                widgets["rem_lbl"].config(fg="#f43f5e")
            else:
                widgets["rem_lbl"].config(fg="#10b981")
                
            # Draw Canvas progress bar
            canvas = widgets["canvas"]
            canvas.delete("all")
            width = canvas.winfo_width()
            if width <= 1:
                width = 240 # Fallback width
                
            color = LICENSE_COLORS.get(lic, "#64748b")
            # Track rect
            canvas.create_rectangle(0, 0, width, 8, fill="#1e293b", outline="", width=0)
            # Fill rect
            fill_w = int(width * (percentage / 100))
            if fill_w > 0:
                canvas.create_rectangle(0, 0, fill_w, 8, fill=color, outline="", width=0)

    # ─── Canvas Charts Rendering ──────────────────────────────────────────────
    def render_charts(self):
        self.render_lic_chart()
        self.render_dept_chart()
        self.render_plant_chart()
        
    def render_lic_chart(self):
        c = self.can_lic
        c.delete("all")
        self.lic_bars_map.clear()
        
        w = c.winfo_width()
        h = c.winfo_height()
        if w <= 1 or h <= 1:
            w, h = 320, 220
            
        pad_x, pad_top, pad_bot = 40, 25, 30
        chart_w = w - pad_x * 2
        chart_h = h - pad_top - pad_bot
        
        # Calculate counts
        counts = {}
        for r in self.filtered_data:
            l = get_row_value(r, LICENSE_KEYS, "Other")
            counts[l] = counts.get(l, 0) + 1
            
        labels = sorted(list(counts.keys()))
        if not labels:
            c.create_text(w/2, h/2, text="No Data Available", fill="#64748b", font=("Helvetica", 10))
            return
            
        max_val = max(counts.values())
        if max_val == 0:
            max_val = 1
            
        # Draw background horizontal grid lines
        grid_lines = 4
        for i in range(grid_lines + 1):
            y_val = pad_top + chart_h - (chart_h * (i / grid_lines))
            c.create_line(pad_x, y_val, w - pad_x, y_val, fill="#1e293b", dash=(3, 3))
            
        bar_count = len(labels)
        bar_gap = 10
        total_gaps_w = bar_gap * (bar_count - 1)
        bar_w = (chart_w - total_gaps_w) / bar_count
        
        for idx, lbl in enumerate(labels):
            val = counts[lbl]
            bar_h = chart_h * (val / max_val)
            bx1 = pad_x + idx * (bar_w + bar_gap)
            by1 = pad_top + chart_h - bar_h
            bx2 = bx1 + bar_w
            by2 = pad_top + chart_h
            
            color = LICENSE_COLORS.get(lbl, LICENSE_COLORS["Other"])
            bar_id = c.create_rectangle(bx1, by1, bx2, by2, fill=color, outline="", width=0)
            
            # Record coordinates for hover tooltip
            self.lic_bars_map[bar_id] = (lbl, val)
            
            # Label
            c.create_text((bx1+bx2)/2, by2 + 12, text=lbl, fill="#94a3b8", font=("Helvetica", 7, "bold"))
            # Value on top of bar
            if bar_h > 15:
                c.create_text((bx1+bx2)/2, by1 - 8, text=str(val), fill="#f8fafc", font=("Helvetica", 8))

    def render_dept_chart(self):
        c = self.can_dept
        c.delete("all")
        self.dept_bars_map.clear()
        
        w = c.winfo_width()
        h = c.winfo_height()
        if w <= 1 or h <= 1:
            w, h = 320, 220
            
        pad_x_left, pad_x_right, pad_y = 65, 30, 20
        chart_w = w - pad_x_left - pad_x_right
        chart_h = h - pad_y * 2
        
        counts = {}
        for r in self.filtered_data:
            d = get_row_value(r, DEPT_KEYS, "Other")
            counts[d] = counts.get(d, 0) + 1
            
        # Sort departments by count desc
        sorted_depts = sorted(counts.items(), key=lambda x: x[1], reverse=True)[:7] # Limit top 7 depts
        if not sorted_depts:
            c.create_text(w/2, h/2, text="No Data Available", fill="#64748b", font=("Helvetica", 10))
            return
            
        max_val = max(counts.values())
        if max_val == 0:
            max_val = 1
            
        bar_count = len(sorted_depts)
        bar_gap = 6
        total_gaps_h = bar_gap * (bar_count - 1)
        bar_h = (chart_h - total_gaps_h) / bar_count
        
        for idx, (dept, val) in enumerate(sorted_depts):
            bar_w = chart_w * (val / max_val)
            bx1 = pad_x_left
            by1 = pad_y + idx * (bar_h + bar_gap)
            bx2 = bx1 + bar_w
            by2 = by1 + bar_h
            
            color = DEPT_COLORS.get(dept, DEPT_COLORS["Other"])
            bar_id = c.create_rectangle(bx1, by1, bx2, by2, fill=color, outline="", width=0)
            
            # Map bar ID
            self.dept_bars_map[bar_id] = (dept, val)
            
            # Label
            c.create_text(bx1 - 8, (by1+by2)/2, text=dept, fill="#94a3b8", font=("Helvetica", 7, "bold"), anchor="e")
            # Value
            c.create_text(bx2 + 10, (by1+by2)/2, text=str(val), fill="#f8fafc", font=("Helvetica", 8), anchor="w")

    def render_plant_chart(self):
        c = self.can_plant
        c.delete("all")
        self.plant_bars_map.clear()
        
        w = c.winfo_width()
        h = c.winfo_height()
        if w <= 1 or h <= 1:
            w, h = 320, 220
            
        pad_x, pad_top, pad_bot = 40, 25, 30
        chart_w = w - pad_x * 2
        chart_h = h - pad_top - pad_bot
        
        # Plant ID prefix is first 4 characters
        counts = {}
        for r in self.filtered_data:
            sapid = get_row_value(r, SAPID_KEYS, "Unknown")
            prefix = sapid[:4] if len(sapid) >= 4 else "Other"
            counts[prefix] = counts.get(prefix, 0) + 1
            
        labels = sorted(list(counts.keys()))[:8] # Limit top 8 prefixes
        if not labels:
            c.create_text(w/2, h/2, text="No Data Available", fill="#64748b", font=("Helvetica", 10))
            return
            
        max_val = max(counts.values())
        if max_val == 0:
            max_val = 1
            
        # Draw background grids
        grid_lines = 4
        for i in range(grid_lines + 1):
            y_val = pad_top + chart_h - (chart_h * (i / grid_lines))
            c.create_line(pad_x, y_val, w - pad_x, y_val, fill="#1e293b", dash=(3, 3))
            
        bar_count = len(labels)
        bar_gap = 10
        total_gaps_w = bar_gap * (bar_count - 1)
        bar_w = (chart_w - total_gaps_w) / bar_count
        
        palette = ['#0ea5e9', '#6366f1', '#a855f7', '#ec4899', '#14b8a6', '#f59e0b', '#10b981']
        
        for idx, lbl in enumerate(labels):
            val = counts[lbl]
            bar_h = chart_h * (val / max_val)
            bx1 = pad_x + idx * (bar_w + bar_gap)
            by1 = pad_top + chart_h - bar_h
            bx2 = bx1 + bar_w
            by2 = pad_top + chart_h
            
            color = palette[idx % len(palette)]
            bar_id = c.create_rectangle(bx1, by1, bx2, by2, fill=color, outline="", width=0, tags="bar")
            
            # Map bar ID to prefix and count details
            self.plant_bars_map[bar_id] = (lbl, val)
            
            # Label
            c.create_text((bx1+bx2)/2, by2 + 12, text=lbl, fill="#94a3b8", font=("Helvetica", 7, "bold"))
            # Value
            if bar_h > 15:
                c.create_text((bx1+bx2)/2, by1 - 8, text=str(val), fill="#f8fafc", font=("Helvetica", 8))

    def on_chart_motion(self, event, canvas, bars_map, tooltip_obj):
        item = canvas.find_withtag("current")
        if item and item[0] in bars_map:
            lbl, val = bars_map[item[0]]
            tooltip_obj.show(event.x, event.y, f"{lbl}: {val}")
        else:
            tooltip_obj.hide()
            
    def on_plant_chart_click(self, event):
        item = self.can_plant.find_withtag("current")
        if item and item[0] in self.plant_bars_map:
            prefix, count = self.plant_bars_map[item[0]]
            # Extract matching rows
            matching_rows = []
            for r in self.filtered_data:
                sapid = get_row_value(r, SAPID_KEYS)
                r_prefix = sapid[:4] if len(sapid) >= 4 else "Other"
                if r_prefix == prefix:
                    matching_rows.append(r)
            # Open Plant ID table modal window
            PlantIDDetailsModal(self, prefix, matching_rows)

    # ─── Explorer Render Methods ──────────────────────────────────────────────
    def refresh_explorer(self):
        selected_cat = self.exp_filter_lic_var.get()
        
        # Calculate group explorer counts
        group_counts = {}
        for r in self.filtered_data:
            # Check explorer category filter
            lic = get_row_value(r, LICENSE_KEYS)
            if selected_cat != "ALL" and lic != selected_cat:
                continue
                
            func = get_row_value(r, FUNC_KEYS, "Unknown")
            if func:
                group_counts[func] = group_counts.get(func, 0) + 1
                
        sorted_groups = sorted(group_counts.keys())
        
        # Group Pagination calculations
        total_pages = math.ceil(len(sorted_groups) / self.explorer_page_size) or 1
        if self.explorer_page > total_pages:
            self.explorer_page = total_pages
        if self.explorer_page < 1:
            self.explorer_page = 1
            
        start_idx = (self.explorer_page - 1) * self.explorer_page_size
        end_idx = min(start_idx + self.explorer_page_size, len(sorted_groups))
        page_groups = sorted_groups[start_idx:end_idx]
        
        # Clear buttons
        for w in self.exp_btn_container.winfo_children():
            w.destroy()
            
        if not page_groups:
            # Empty state
            lbl_empty = tk.Label(self.exp_btn_container, text="No groups found", font=("Helvetica", 9), bg="#161e31", fg="#64748b")
            lbl_empty.pack(fill="both", expand=True, pady=30)
            self.exp_lbl_table_title.config(text="Assigned Users")
            self.exp_lbl_table_sub.config(text="No active group selected")
            self.exp_tree.delete(*self.exp_tree.get_children())
            return
            
        # Select first available if selection invalid
        if not self.explorer_selected_func or self.explorer_selected_func not in sorted_groups:
            self.explorer_selected_func = page_groups[0]
            
        for g_name in page_groups:
            count = group_counts[g_name]
            # Custom styled toggle list buttons
            is_active = (g_name == self.explorer_selected_func)
            bg_color = "#0ea5e9" if is_active else "#1e293b"
            fg_color = "#0b0f19" if is_active else "#f8fafc"
            
            btn = tk.Button(
                self.exp_btn_container,
                text=f"{g_name} ({count})",
                font=("Helvetica", 8, "bold"),
                bg=bg_color, fg=fg_color,
                activebackground=bg_color, activeforeground=fg_color,
                bd=0, anchor="w", padx=12, pady=6, cursor="hand2",
                command=lambda gn=g_name: self.set_explorer_group(gn)
            )
            btn.pack(fill="x", pady=2)
            
        # Pagination Controls row at the bottom of buttons list
        pag_row = tk.Frame(self.exp_btn_container, bg="#161e31", pady=8)
        pag_row.pack(fill="x", side="bottom")
        
        btn_prev = tk.Button(pag_row, text="◀", font=("Helvetica", 8, "bold"), bg="#1e293b", fg="#f8fafc", activebackground="#334155", activeforeground="#f8fafc", bd=0, padx=6, pady=2, cursor="hand2", command=self.explorer_prev_page)
        btn_prev.pack(side="left")
        if self.explorer_page == 1:
            btn_prev.config(state="disabled")
            
        lbl_info = tk.Label(pag_row, text=f"Page {self.explorer_page} of {total_pages}", font=("Helvetica", 8), bg="#161e31", fg="#94a3b8")
        lbl_info.pack(side="left", fill="x", expand=True)
        
        btn_next = tk.Button(pag_row, text="▶", font=("Helvetica", 8, "bold"), bg="#1e293b", fg="#f8fafc", activebackground="#334155", activeforeground="#f8fafc", bd=0, padx=6, pady=2, cursor="hand2", command=self.explorer_next_page)
        btn_next.pack(side="right")
        if self.explorer_page == total_pages:
            btn_next.config(state="disabled")
            
        self.render_explorer_table(self.explorer_selected_func, selected_cat)

    def set_explorer_group(self, group_name):
        self.explorer_selected_func = group_name
        self.refresh_explorer()
        
    def explorer_prev_page(self):
        if self.explorer_page > 1:
            self.explorer_page -= 1
            self.refresh_explorer()
            
    def explorer_next_page(self):
        self.explorer_page += 1
        self.refresh_explorer()

    def render_explorer_table(self, group_name, selected_cat):
        # Update Table headings
        self.exp_lbl_table_title.config(text=group_name)
        
        # Extract matching users
        matching_users = []
        for r in self.filtered_data:
            # Check explorer category filter
            lic = get_row_value(r, LICENSE_KEYS)
            if selected_cat != "ALL" and lic != selected_cat:
                continue
                
            func = get_row_value(r, FUNC_KEYS)
            if func == group_name:
                matching_users.append(r)
                
        self.exp_lbl_table_sub.config(text=f"Unique Assigned Persons: {len(matching_users)}")
        
        self.exp_tree.delete(*self.exp_tree.get_children())
        
        # Sort matching users by department
        matching_users = sorted(matching_users, key=lambda x: get_row_value(x, DEPT_KEYS))
        
        for u in matching_users:
            self.exp_tree.insert("", "end", values=(
                get_row_value(u, NAME_KEYS),
                get_row_value(u, LICENSE_KEYS),
                get_row_value(u, DEPT_KEYS)
            ))

    def on_explorer_user_double_click(self, event):
        item = self.exp_tree.selection()
        if not item:
            return
        vals = self.exp_tree.item(item[0], "values")
        name = vals[0]
        # Match user from current raw list
        for u in self.raw_data:
            if get_row_value(u, NAME_KEYS) == name:
                UserDetailModal(self, u)
                break

    # ─── Employee Ledger Tab Rendering ────────────────────────────────────────
    def refresh_ledger_view(self):
        search_term = self.ent_search.get().strip().lower()
        lic_filter = self.ledg_filter_lic_var.get()
        dept_filter = self.ledg_filter_dept_var.get()
        
        # Filter raw rows list
        list_rows = []
        for r in self.raw_data:
            match_lic = (lic_filter == "ALL" or get_row_value(r, LICENSE_KEYS) == lic_filter)
            match_dept = (dept_filter == "ALL" or get_row_value(r, DEPT_KEYS) == dept_filter)
            
            sapid = get_row_value(r, SAPID_KEYS)
            name = get_row_value(r, NAME_KEYS)
            match_search = (not search_term or search_term in sapid.lower() or search_term in name.lower())
            
            if match_lic and match_dept and match_search:
                list_rows.append({
                    "sapid": sapid,
                    "name": name,
                    "license": get_row_value(r, LICENSE_KEYS),
                    "department": get_row_value(r, DEPT_KEYS),
                    "group": get_row_value(r, GROUP_KEYS),
                    "lastlogon": get_row_value(r, LOGON_KEYS)
                })
                
        # Sort elements
        def get_sort_val(x):
            return x.get(self.ledger_sort_col, "").lower()
            
        list_rows = sorted(list_rows, key=get_sort_val, reverse=not self.ledger_sort_asc)
        
        total_items = len(list_rows)
        total_pages = math.ceil(total_items / self.ledger_page_size) or 1
        
        if self.ledger_page > total_pages:
            self.ledger_page = total_pages
        if self.ledger_page < 1:
            self.ledger_page = 1
            
        start_idx = (self.ledger_page - 1) * self.ledger_page_size
        end_idx = min(start_idx + self.ledger_page_size, total_items)
        page_items = list_rows[start_idx:end_idx]
        
        # Truncate and rewrite Treeview rows
        self.tree_ledg.delete(*self.tree_ledg.get_children())
        
        for item in page_items:
            self.tree_ledg.insert("", "end", values=(
                item["sapid"],
                item["name"],
                item["license"],
                item["department"],
                item["group"],
                item["lastlogon"]
            ))
            
        # Update pagination info labels
        if total_items == 0:
            self.lbl_pag_info.config(text="Showing 0 to 0 of 0 employees")
            self.btn_pag_prev.config(state="disabled")
            self.btn_pag_next.config(state="disabled")
        else:
            self.lbl_pag_info.config(text=f"Showing {start_idx + 1} to {end_idx} of {total_items} employees")
            self.btn_pag_prev.config(state="normal" if self.ledger_page > 1 else "disabled")
            self.btn_pag_next.config(state="normal" if self.ledger_page < total_pages else "disabled")

    def sort_ledger(self, column):
        if self.ledger_sort_col == column:
            self.ledger_sort_asc = not self.ledger_sort_asc
        else:
            self.ledger_sort_col = column
            self.ledger_sort_asc = True
        self.ledger_page = 1
        self.refresh_ledger_view()

    def ledger_prev_page(self):
        if self.ledger_page > 1:
            self.ledger_page -= 1
            self.refresh_ledger_view()
            
    def ledger_next_page(self):
        self.ledger_page += 1
        self.refresh_ledger_view()

    def on_ledger_row_double_click(self, event):
        item = self.tree_ledg.selection()
        if not item:
            return
        vals = self.tree_ledg.item(item[0], "values")
        sapid = vals[0]
        # Open user detailed modal window
        for u in self.raw_data:
            if get_row_value(u, SAPID_KEYS) == sapid:
                UserDetailModal(self, u)
                break

    # ─── Config Backup Import / Export ────────────────────────────────────────
    def export_config(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON File", "*.json")],
            initialfile=f"sap_license_config_{datetime.date.today().isoformat()}.json"
        )
        if not file_path:
            return
        try:
            config = {
                "dashboard_sheet_url": self.active_url,
                "cached_sheet_data": self.raw_data,
                "limits": self.limits,
                "exported_at": datetime.datetime.now().isoformat(),
                "app": "sapid-license-analyser-Desktop",
                "version": "1.0.0"
            }
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(config, f, indent=2)
            self.show_toast("Configuration exported successfully!")
        except Exception as e:
            append_log("ERROR", f"Export configuration file failed: {e}")
            messagebox.showerror("Export Failed", f"Failed to save JSON config:\n{e}")

    def import_config(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("JSON Config File", "*.json")]
        )
        if not file_path:
            return
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                config = json.load(f)
                
            if "cached_sheet_data" not in config:
                raise Exception("JSON file does not contain valid cached spreadsheet data.")
                
            self.active_url = config.get("dashboard_sheet_url", "")
            self.raw_data = config.get("cached_sheet_data", [])
            self.limits = config.get("limits", DEFAULT_LIMITS.copy())
            
            # Write to active local configs
            self.save_config()
            
            # Sync URL Box Entry Text
            self.ent_url.delete(0, "end")
            self.ent_url.insert(0, self.active_url)
            
            # Refresh layout
            self.refresh_ui()
            self.show_toast("Configuration imported successfully!")
        except Exception as e:
            append_log("ERROR", f"Import config file failed: {e}")
            messagebox.showerror("Import Failed", f"Failed to load JSON config:\n{e}")

    # ─── Modal Show Wrappers ──────────────────────────────────────────────────
    def show_user_manual(self):
        TextModal(self, "Help & Documentation", "User Manual & Guidelines", USER_MANUAL_TEXT.strip())
        
    def show_privacy_policy(self):
        TextModal(self, "Security & Privacy Policy", "Application Privacy Policy", PRIVACY_TEXT.strip())
        
    def show_licenses(self):
        TextModal(self, "Third-Party Notices", "Open Source Software Licenses", LICENSES_TEXT.strip())
        
    def show_logs_modal(self):
        ErrorLogsModal(self)

    # ─── Styled Toast Alert Notification overlay ──────────────────────────────
    def show_toast(self, message):
        toast = tk.Toplevel(self)
        toast.overrideredirect(True)
        toast.attributes("-topmost", True)
        
        # Center toast relative to master window
        mx = self.winfo_x() + (self.winfo_width() / 2) - 130
        my = self.winfo_y() + 80
        toast.geometry(f"280x45+{int(mx)}+{int(my)}")
        
        frame = tk.Frame(toast, bg="#0ea5e9", highlightbackground="#ffffff", highlightthickness=1)
        frame.pack(fill="both", expand=True)
        
        lbl = tk.Label(frame, text=f"✅  {message}", font=("Helvetica", 9, "bold"), bg="#0ea5e9", fg="#0b0f19")
        lbl.pack(fill="both", expand=True)
        
        # Auto self-destruct toast window overlay after 2 seconds
        self.after(2000, toast.destroy)

    # ─── Frame Resizing Event Handler ──────────────────────────────────────────
    def on_window_resize(self, event):
        # Force Canvas chart redraws on window configure events
        # Check if geometry changes are significant to avoid recursive loops
        if self.current_tab == "dashboard":
            # Redraw graphs
            self.render_charts()
            if hasattr(self, "allocated_counts"):
                self.render_quota_sidebar(self.allocated_counts)

# ─── Main Program Execution ───────────────────────────────────────────────────
if __name__ == "__main__":
    try:
        main_app = SAPLicenseAnalyserApp()
        # Bind root resize configurations
        main_app.bind("<Configure>", lambda e: main_app.on_window_resize(e) if e.widget == main_app else None)
        main_app.mainloop()
    except Exception as e:
        append_log("FATAL", f"App crash on startup: {e}")
        raise e
